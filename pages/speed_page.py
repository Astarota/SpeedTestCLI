import io
import time
from openpyxl import Workbook
import openpyxl as O
import speedtest
import re
import sys
import psycopg2
from psycopg2 import Error
from datetime import datetime
import hashlib
from .connection import Connection

cn=Connection()

class SpeedPage():
	def asking_name(self):
		print("Введите ваше ИМЯ:")
		author=input()
		return author
	

	def asking_serial_number(self):
		print("Введите ваш Серийный номер:")
		s=input()
		serialno=s.replace("S/N: ","")
		return serialno
	
	def id_of_test(self):
		try:
			connection = psycopg2.connect(user="dssadmin",
				password="dssadmin",
				host="10.10.2.180",
				port="5432",
				database="devices")
			
			cursor = connection.cursor()
			cursor.execute(""" SELECT MAX(testId) FROM testhistory """)
			result = cursor.fetchall()		
			for i in result:
				testId = i[0] + 1

		
		except (Exception, Error) as error:
			print("Ошибка при работе с PostgreSQL", error)
		finally:
			if connection:
				cursor.close()
				connection.close()
		return testId

	def port_number(self):
		print("Какой это порт:")
		port_number=input()
		return port_number


	def speedtest_database(self,author,serialno,testId,pretty_serialno,port_number,j):
		for i in range(1,3):
				current_date = datetime.now().strftime("%Y-%m-%d")
				current_time = datetime.now().strftime("%H:%M:%S")
				start_time=datetime.now()
				st=speedtest.Speedtest(secure=True)
				st.get_best_server()
				download_number=st.download()
				print("Ваша входящяя скорость:", end=' ')
				download_speed=self.test_download_test(download_number)
				print(download_speed)
				upload_number=st.upload()
				print("Ваша исходящяя скорость:", end=' ')
				upload_speed=self.test_upload_test(upload_number)
				print(upload_speed)
				stop_time=datetime.now()
				duration=str((stop_time - start_time).seconds)+' сек'
				self.excel_uploading(download_speed,upload_speed,j,pretty_serialno,port_number)
				print("Количество времени потраченная на тест:", end=' ')
				print(duration)
				print("Тест номер:", end=' ')
				print(testId)
				cn.connect_to_database(author,serialno,testId,download_speed,upload_speed,
					duration,current_date,current_time,port_number)
			

	def pretty_speed(self,speed):
		unit = 'bps'
		kmg = ['', 'K', 'M', 'G']
		i = 0
		while speed >= 1000:
			speed /= 1000
			i += 1
		return "{:.2f}".format(speed) + ' ' + kmg[i] + unit

	def speed_measure(self,speed):
		i=0
		while speed >= 1000:
			speed /= 1000
			i+=1
		return speed

	def pretty_file_format(self,serialno):
		encoding='.xlsx'
		return serialno+encoding

	def test_download_test(self,download_number):
		download_speed=self.speed_measure(download_number)
		download_beauty_speed = self.pretty_speed(download_number)
		if(download_speed<100):
			print("FAIL", end=' ')
		return download_beauty_speed

	def test_upload_test(self,upload_number):
		upload_speed=self.speed_measure(upload_number)
		upload_beauty_speed = self.pretty_speed(upload_number)
		if(upload_speed<100):
			print("FAIL", end=' ')
		return upload_beauty_speed

	def create_excel_file(self,serialno):
		wb=Workbook()
		ws=wb.active
		wb.save(filename=serialno)	


	def excel_uploading(self,download_speed,upload_speed,i,serialno,port_number):
		Excel_file=serialno
		Excel_worksheet="Sheet"
		i=i+1
		wb=O.load_workbook(Excel_file)
		ws=wb[Excel_worksheet]
		ws.cell(i,1).value=port_number
		ws.cell(1,1).value='Port Number'
		ws.cell(1,2).value='Download Speed'
		ws.cell(1,3).value='Upload Speed'
		ws.cell(i,2).value=download_speed
		ws.cell(i,3).value=upload_speed
		wb.save(Excel_file)
		wb.close()
		